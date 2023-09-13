from selenium import webdriver
from shutil import which
from bs4 import BeautifulSoup
import re
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import TimeoutException
import time  
import pandas as pd
import requests
import wget
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

def extract_likes_and_comments(text):
    # Regular expressions to extract likes and comments
    pattern_likes = r'(\d+) likes'
    pattern_comments = r'(\d+) comments'

    match_likes = re.search(pattern_likes, text)
    likes = match_likes.group(1) if match_likes else None

    match_comments = re.search(pattern_comments, text)
    comments = match_comments.group(1) if match_comments else None

    return [likes, comments]

# Set the path for the ChromeDriver extension
chrome_path = which('/usr/local/bin/chromedriver')

# Create a ChromeDriver service
service = Service(executable_path=chrome_path)
driver = webdriver.Chrome(service=service)

# Open Instagram
driver.get('https://www.instagram.com/')
driver.implicitly_wait(20)

# Target username and password input fields
username = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='username']")))
password = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='password']")))

# Enter username and password
username.clear()
username.send_keys("pavanbalu49")
password.clear()
input_password = input("Password: ")
password.send_keys(input_password)

# Target the login button and click it
button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))).click()

# Handle alerts if present
time.sleep(5)
try:
    alert = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[contains(text(), "Not Now")]'))).click()
except (NoSuchElementException, ElementClickInterceptedException):
    pass

time.sleep(5)

try:
    alert2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "Not Now")]'))).click()
except (NoSuchElementException, ElementClickInterceptedException, TimeoutException):
    pass

# Create a new Excel workbook
workbook = Workbook()
time.sleep(4)

# Initialize variables
start = time.time()
current_datetime = datetime.datetime.now()
formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
filename = f"Instagram_{formatted_datetime}.xlsx"
activeSheetNotUsed = True

# List of keywords to search for
keywordsList = [
    "#STEMPUSH",
    "#STEMPUSH(PathwaysforUnderrepresentedStudentstoHigherEducation)Network",
    "#STEMPUSHNetwork",
    "#STEMPUSHAlliance",
    "#STEMPUSHNetworkHub",
    "#STEMPUSHNetworkAlliance",
    "#STEMPUSHNetwork,UniversityofPittsburgh",
    "#STEMPUSHAlliance,UniversityofPittsburgh",
    "#NSFINCLUDESAlliance1930990TheSTEMPUSHNetwork",
    "#NSFINCLUDESAlliance1930990",
    "#PathwaysforUnderrepresentedStudentstoHigherEducation",
    "#PreCollegeSTEMPrograms(PCSPs)"
]

for keyword in keywordsList:
    # Target the search input field
    search_icon = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@aria-label='Search']"))).click()
    time.sleep(5)
    searchbox = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Search']")))
    time.sleep(4)

    # Search for the keyword
    searchbox.clear()
    time.sleep(2)
    searchbox.send_keys(keyword)
    time.sleep(10)

    header_labels = ["Username", "Description", "Likes", "Comments"]

    if activeSheetNotUsed:
        sheet = workbook.active
        sheet.title = keyword + " Data"
        activeSheetNotUsed = False
    else:
        sheet = workbook.create_sheet(keyword + ' Data')

    # Set header labels and formatting
    for col_idx, header_label in enumerate(header_labels, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header_label
        cell.font = Font(bold=True, size=12)
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 130

    found_search = True
    try:
        found_search = driver.find_element(By.XPATH, '(//div[@class = "x9f619 xjbqb8w x78zum5 x168nmei x13lgxp2 x5pf9jr xo71vjh xxbr6pl xbbxn1n xwib8y2 x1y1aw1k x1uhb9sk x1plvlek xryxfnj x1c4vz4f x2lah0s xdt5ytf xqjyukv x1qjc9v5 x1oa3qoh x1nhvcw1"])[1]')
    except NoSuchElementException:
        found_search = False
        pass

    if not found_search:
        driver.find_element(By.XPATH, "//*[@aria-label='Search']").click()
        time.sleep(5)
        print('Found 0 links to images for the word ', keyword)
        workbook.save(filename)
        continue

    searchbox.send_keys(Keys.ENTER)
    time.sleep(5)
    searchbox.send_keys(Keys.ENTER)
    time.sleep(15)

    n_scrolls = 2
    for j in range(0, n_scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)

    # Find links to images
    anchors = driver.find_elements(By.XPATH, "//a[contains(@href, '/p/')]")
    anchors = [a.get_attribute('href') for a in anchors]
    anchors = [a for a in anchors]
    driver.find_element(By.XPATH, "//*[@aria-label='Search']").click()
    print('Found ' + str(len(anchors)) + ' links to images for the word ', keyword)
    print("\n" * 6)

    # Iterate through image links and extract data
    for link in anchors:
        response = requests.get(link)
        html_content = response.content
        # Create a BeautifulSoup object with the HTML content
        soup = BeautifulSoup(html_content, 'html.parser')

        # Find the meta tag with the specified attributes
        meta_tag_for_name = soup.find('meta', attrs={'name': 'twitter:title'})
        meta_tag_for_desc = soup.find('meta', attrs={'property': 'og:title'})
        meta_tag_for_additional = soup.find('meta', attrs={'name': 'description'})
        likes, comments = extract_likes_and_comments(meta_tag_for_additional.get('content'))
        # Extract the content attribute value
        username = meta_tag_for_name.get('content')
        username = username.split("•")[0].strip()
        desc = meta_tag_for_desc.get('content')

        # Print the extracted content
        print('username-->', username)
        print('description-->', desc)
        print('likes-->', likes, 'comments-->', comments)
        print('----------------------------------------------------------------')

        # Append data to the Excel sheet
        sheet.append([username, desc, int(likes), int(comments)])
        workbook.save(filename)

# Close the ChromeDriver
driver.quit()
