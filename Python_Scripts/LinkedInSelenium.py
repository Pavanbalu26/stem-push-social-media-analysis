# Import necessary libraries
from selenium import webdriver
from shutil import which
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException, WebDriverException
import time
import pandas as pd
import wget
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# Function to extract integers from a string
def extract_integers(string):
    result = ""
    for char in string:
        if char.isdigit():
            result += char
    return int(result)

# Set the path for the ChromeDriver extension
chrome_path = which('/usr/local/bin/chromedriver')

# Create a ChromeDriver service
service = Service(executable_path=chrome_path)
driver = webdriver.Chrome(service=service)

# Open LinkedIn
driver.get('https://www.linkedin.com/')
driver.implicitly_wait(10)

# Target username and password input fields
username = WebDriverWait(driver, 25).until(EC.element_to_be_clickable((By.XPATH, "//*[@name='session_key']")))
password = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[@name='session_password']")))

# Enter username and password
username.clear()
username.send_keys("pavanbalu.korlepara@gmail.com")
password.clear()
input_password = input("Password: ")
password.send_keys(input_password)

# Click the login button
button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@type='submit']"))).click()

time.sleep(5)

# Creating a Workbook
workbook = Workbook()
time.sleep(4)
start = time.time()
current_datetime = datetime.datetime.now()
formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
filename = f"LinkedIn_{formatted_datetime}.xlsx"
activeSheetNotUsed = True

# List of keywords to search for
keywordsList = ['Stem Education', 'Stem Push']

for keyword in keywordsList:
    # Target the search input field
    searchBox = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='Search']")))
    searchBox.clear()
    searchBox.send_keys(keyword)
    time.sleep(2)
    searchBox.send_keys(Keys.ENTER)
    time.sleep(4)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Posts')]")))
    Posts_button = driver.find_element(By.XPATH, "//button[contains(., 'Posts')]")
    Posts_button.click()

    n_scrolls = 1
    for j in range(0, n_scrolls):
        time.sleep(10)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        print(j, " Page is scrolled")
    posts = driver.find_elements(By.XPATH, "//div[@class='pt1 mb2 artdeco-card']")

    print('Parsing Begins')
    time.sleep(4)

    if activeSheetNotUsed:
        sheet = workbook.active
        sheet.title = keyword + " Data"
        activeSheetNotUsed = False
    else:
        sheet = workbook.create_sheet(keyword + ' Data')
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 130
    sheet.column_dimensions['C'].width = 40

    header_labels = ["Usernames", "Descriptions", "HashTags", "Reactions", "Comments", "Reposts"]
    
    # Set header labels and formatting
    for col_idx, header_label in enumerate(header_labels, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header_label
        cell.font = Font(bold=True, size=12)

    for post in posts:
        username = ''
        description = ''
        hashtags = ''
        reactions = 0
        comments = 0
        reposts = 0

        usernames = post.find_elements(By.XPATH, ".//span[contains(@class,'update-components-actor__name t-14 t-bold hoverable-link-text')]//span[@dir = 'ltr']")
        
        # Extract usernames
        for nth_time in range(len(usernames)):
            if nth_time != len(usernames) - 1:
                username = username + usernames[nth_time].text + ','
            else:
                username = username + usernames[nth_time].text
        
        try:
            descriptions = post.find_elements(By.XPATH, ".//span[@class='break-words']")
            
            # Extract descriptions
            for nth_time in range(len(descriptions)):
                if nth_time != len(descriptions) - 1:
                    description = description + descriptions[nth_time].text + ','
                else:
                    description = description + descriptions[nth_time].text
                    
        except (NoSuchElementException, WebDriverException):
            pass

        try:
            anchorTags = post.find_elements(By.XPATH, ".//span[@class='break-words']//a")
            
            # Extract hashtags
            for nth_time in range(len(anchorTags)):
                if nth_time != len(anchorTags) - 1:
                    hashtags = hashtags + anchorTags[nth_time].get_attribute("textContent") + ','
                else:
                    hashtags = hashtags + anchorTags[nth_time].get_attribute("textContent")
                    
        except (NoSuchElementException, WebDriverException):
            pass

        try:
            reactions = post.find_element(By.XPATH, ".//span[@class='social-details-social-counts__reactions-count']").text
            
        except (NoSuchElementException, WebDriverException):
            pass

        try:
            comments = post.find_element(By.XPATH, ".//button[contains(@class,'t-black--light social-details-social-counts__')]").text
            
        except (NoSuchElementException, WebDriverException):
            pass

        try:
            reposts = post.find_element(By.XPATH, ".//button[contains(@class,'ember-view t-black--light')]").text
            
        except (NoSuchElementException, WebDriverException):
            pass

        # Extract integers from reactions, comments, and reposts
        reactions = extract_integers(reactions) if reactions != 0 else 0
        comments = extract_integers(comments) if comments != 0 else 0
        reposts = extract_integers(reposts) if reposts != 0 else 0

        sheet.append([username, description, hashtags, reactions, comments, reposts])
        workbook.save(filename)

    font_size = 12  # Adjust the size as desired
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(size=font_size)
    workbook.save(filename)
